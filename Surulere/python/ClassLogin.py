"""
CodeLagos Project by:
Anthony Oluwasunhan
08025100553
tony.oluwasunhan@gmail.com
Surulere Centre
This project is done and written to solve the problem of students signing in for another student in
CodeLagos Class. With this solution students can not have access to the class without their username and code.
The algorithm is written for project purpose, the details are automatically stored in an excel sheet
and information are retrieved from same excel sheet.
This code can be further improved than what is submitted here.
"""

#required for excel
import openpyxl

#required to get the current working directory
import os

#required to format the excel sheet
from openpyxl.styles import Font


#Get the details of student
full_name = input("Enter First and Surname: ")
phone = input("Enter your Phone number ")
email = input("Enter your Email address ")
course = input("Enter Course registered for ")
print()
print('CREATE A USERNAME AND PINCODE')
user_name = input('Username: \n')
pincode = input("Pincode \n")

wb = openpyxl.Workbook()
wb.create_sheet(index=0,title ='codelagos')
wb.create_sheet(index=1,title ='details')
codelagos = wb.get_sheet_by_name('codelagos')
details = wb.get_sheet_by_name('details')

row = codelagos.max_row + 1

#Assigns the Header in column A1 and B1
codelagos['A1'] = 'USERNAME'
codelagos['B1'] = 'PINCODE'

#formats the header
fontObj1 = Font(bold=True)
codelagos['A1'].font = fontObj1
codelagos['B1'].font = fontObj1

#stores infromation on the first sheet
codelagos['A'+ str(row)] = user_name
codelagos['B' + str(row)] = pincode

##Assigns the Header in columns on the second sheet of the workbook
details['A1'] = 'FULLNAME'
details['B1'] = 'PINCODE'
details['C1'] = 'PHONE'
details['D1'] = 'EMAIL'
details['E1'] = 'CLASS'

#formats the header
fontObj1 = Font(bold=True)
details['A1'].font = fontObj1
details['B1'].font = fontObj1
details['C1'].font = fontObj1
details['D1'].font = fontObj1
details['E1'].font = fontObj1

#stores infromation on the second sheet
details['A'+ str(row)] = full_name
details['B' + str(row)] = pincode
details['C'+ str(row)] = phone
details['D'+ str(row)] = email
details['E'+ str(row)] = course

#saves all information to this file
wb.save('LoginData.xlsx')

print('Congratulations {}, Registration Successful'.format(full_name).upper())


#Log in after user have registered
print()
print('Now Login to access CodeLagos Class')
log_in1 = input("Enter your Username ")

#Get your pincode from the excel sheet
retrieve_user = codelagos['A2'].value
retrieve_pin = codelagos['B2'].value

#compares the username entered from the one retrived from the excel sheet
if log_in1 != retrieve_user:
    print('Unknown User')
else:
    #compares the pincode entered from the one retrived from the excel sheet
    log_in2 = input("Enter your Pincode ")
    if log_in2 == retrieve_pin:
        print()
        print('{}, you are welcome to {} Class of CodeLagos'.format(full_name,course).upper())
        #gets the current working directory where the excel file is stored
        cwd = os.getcwd()
        print()
        print('Open the excel sheet LoginData.xlsx in the location below to see saved file')
        print(cwd)
        print('Thank you')
    else:
        print('Wrong Pin')
print('Code by: Anthony Oluwasunhan, 08025100553, tony.oluwasunhan@gmail.com')
